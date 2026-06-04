"""Image metadata export service.

Exports per-image metadata (persons, date, location, GPS) to CSV or XLSX.
Designed for archival, research, and family-history use cases.
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional, Set

from sqlalchemy.orm import Session

from app.db.models import Face, Image, Person

log = logging.getLogger(__name__)

UNKNOWN = "Ismeretlen"

# Field key constants — use these when specifying which columns to include.
FIELD_FILENAME = "filename"
FIELD_RELPATH = "relative_path"
FIELD_PERSONS = "persons"
FIELD_DATE = "date"
FIELD_LOCATION = "location"
FIELD_GPS = "gps"

COL_FILENAME = "Filename"
COL_PATH = "Path"
COL_PERSONS = "Persons"
COL_DATE = "Date"
COL_LOCATION = "Location"
COL_LAT = "Lat"
COL_LON = "Lon"

ALL_FIELDS: Set[str] = {
    FIELD_FILENAME,
    FIELD_RELPATH,
    FIELD_PERSONS,
    FIELD_DATE,
    FIELD_LOCATION,
    FIELD_GPS,
}

PERSON_MODE_LIST = "list"  # all persons comma-separated in one column
PERSON_MODE_COLS = "cols"  # separate Person1, Person2, … columns


@dataclass
class ImageExportRow:
    filename: str
    relative_path: str
    persons: list[str]
    date: str
    location: str
    latitude: Optional[float]
    longitude: Optional[float]


class ImageMetadataExportService:
    """Export per-image metadata (persons, date, location, GPS) to CSV or XLSX.

    Args:
        session: SQLAlchemy session.
    """

    def __init__(self, session: Session) -> None:
        self._session = session

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export_csv(
        self,
        target_path: str | Path,
        fields: Set[str],
        person_mode: str = PERSON_MODE_LIST,
        progress_cb: Optional[Callable[[int, int], None]] = None,
    ) -> Path:
        """Write image metadata to a CSV file.

        Args:
            target_path: Destination ``.csv`` path.
            fields: Set of FIELD_* constants to include.
            person_mode: ``"list"`` (comma-separated) or ``"cols"`` (separate columns).
            progress_cb: Optional ``(done, total)`` callback for progress reporting.

        Returns:
            Path to the written file.
        """
        out = _with_suffix(target_path, ".csv")
        out.parent.mkdir(parents=True, exist_ok=True)

        images = self._all_images()
        total = len(images)
        rows, max_persons = self._build_rows(images, fields, person_mode, progress_cb, total)
        headers = self._build_headers(fields, person_mode, max_persons)

        # utf-8-sig adds a BOM so Excel auto-detects UTF-8 and renders accents correctly.
        with open(out, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=headers, extrasaction="ignore", restval="")
            writer.writeheader()
            writer.writerows(rows)

        log.info("Image metadata CSV: %d row(s) → %s", len(rows), out)
        return out

    def export_xlsx(
        self,
        target_path: str | Path,
        fields: Set[str],
        person_mode: str = PERSON_MODE_LIST,
        progress_cb: Optional[Callable[[int, int], None]] = None,
    ) -> Path:
        """Write image metadata to an XLSX file.

        Requires ``openpyxl`` (``pip install openpyxl``).

        Args:
            target_path: Destination ``.xlsx`` path.
            fields: Set of FIELD_* constants to include.
            person_mode: ``"list"`` or ``"cols"``.
            progress_cb: Optional ``(done, total)`` callback.

        Returns:
            Path to the written file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for XLSX export. "
                "Install it with: pip install openpyxl"
            ) from exc

        out = _with_suffix(target_path, ".xlsx")
        out.parent.mkdir(parents=True, exist_ok=True)

        images = self._all_images()
        total = len(images)
        rows, max_persons = self._build_rows(images, fields, person_mode, progress_cb, total)
        headers = self._build_headers(fields, person_mode, max_persons)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Image Metadata"

        header_fill = PatternFill(start_color="2F4F7F", end_color="2F4F7F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        gps_headers = {COL_LAT, COL_LON}

        for row_dict in rows:
            data = []
            for h in headers:
                val = row_dict.get(h, "")
                if h in gps_headers and val != "":
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = ""
                data.append(val)
            ws.append(data)

        col_widths: dict[int, int] = {}
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is not None:
                    col_widths[cell.column] = max(
                        col_widths.get(cell.column, 0),
                        min(len(str(cell.value)), 60),
                    )
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 4

        ws.freeze_panes = "A2"

        wb.save(out)
        log.info("Image metadata XLSX: %d row(s) → %s", len(rows), out)
        return out

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _all_images(self) -> list[Image]:
        return self._session.query(Image).order_by(Image.id).all()

    def _build_rows(
        self,
        images: list[Image],
        fields: Set[str],
        person_mode: str,
        progress_cb: Optional[Callable[[int, int], None]],
        total: int,
    ) -> tuple[list[dict], int]:
        rows: list[dict] = []
        max_persons = 0
        for idx, image in enumerate(images):
            try:
                export_row = self._build_image_row(image)
                row_dict = self._row_to_dict(export_row, fields, person_mode)
                if person_mode == PERSON_MODE_COLS:
                    max_persons = max(max_persons, len(export_row.persons))
                rows.append(row_dict)
            except Exception:
                log.exception("Error building export row for image id=%s", image.id)
            if progress_cb:
                progress_cb(idx + 1, total)
        return rows, max_persons

    def _build_image_row(self, image: Image) -> ImageExportRow:
        return ImageExportRow(
            filename=Path(image.file_path).name,
            relative_path=image.relative_path or image.file_path,
            persons=self._get_person_names(image),
            date=self._resolve_date(image),
            location=self._resolve_location(image),
            latitude=self._resolve_gps(image)[0],
            longitude=self._resolve_gps(image)[1],
        )

    def _get_person_names(self, image: Image) -> list[str]:
        faces = (
            self._session.query(Face)
            .filter(
                Face.image_id == image.id,
                Face.person_id.isnot(None),
                Face.is_excluded.is_(False),
            )
            .all()
        )
        seen: set[int] = set()
        names: list[str] = []
        for face in faces:
            if face.person_id not in seen:
                seen.add(face.person_id)
                person = self._session.get(Person, face.person_id)
                if person:
                    names.append(person.name)
        return names

    def _resolve_gps(self, image: Image) -> tuple[Optional[float], Optional[float]]:
        if image.image_latitude is not None and image.image_longitude is not None:
            return image.image_latitude, image.image_longitude
        if image.exif_latitude is not None and image.exif_longitude is not None:
            return image.exif_latitude, image.exif_longitude
        if image.place and image.place.latitude is not None:
            return image.place.latitude, image.place.longitude
        return None, None

    def _resolve_date(self, image: Image) -> str:
        if image.photo_date:
            return image.photo_date
        name = Path(image.file_path).stem
        m = re.search(r"(\d{4})[_\-](\d{2})[_\-](\d{2})", name)
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        m = re.search(r"(\d{8})", name)
        if m:
            s = m.group(1)
            return f"{s[:4]}-{s[4:6]}-{s[6:]}"
        m = re.search(r"(\d{4})", name)
        if m:
            return m.group(1)
        return UNKNOWN

    def _resolve_location(self, image: Image) -> str:
        if image.place and not image.place.is_anonymous:
            return image.place.name
        return ""

    def _row_to_dict(
        self,
        row: ImageExportRow,
        fields: Set[str],
        person_mode: str,
    ) -> dict:
        d: dict = {}
        if FIELD_FILENAME in fields:
            d[COL_FILENAME] = row.filename
        if FIELD_RELPATH in fields:
            d[COL_PATH] = row.relative_path
        if FIELD_DATE in fields:
            d[COL_DATE] = row.date
        if FIELD_LOCATION in fields:
            d[COL_LOCATION] = row.location
        if FIELD_GPS in fields:
            d[COL_LAT] = row.latitude if row.latitude is not None else ""
            d[COL_LON] = row.longitude if row.longitude is not None else ""
        if FIELD_PERSONS in fields:
            if person_mode == PERSON_MODE_LIST:
                d[COL_PERSONS] = ", ".join(row.persons)
            else:
                for i, name in enumerate(row.persons, start=1):
                    d[f"Person{i}"] = name
        return d

    def _build_headers(
        self,
        fields: Set[str],
        person_mode: str,
        max_persons: int,
    ) -> list[str]:
        headers: list[str] = []
        if FIELD_FILENAME in fields:
            headers.append(COL_FILENAME)
        if FIELD_RELPATH in fields:
            headers.append(COL_PATH)
        if FIELD_DATE in fields:
            headers.append(COL_DATE)
        if FIELD_LOCATION in fields:
            headers.append(COL_LOCATION)
        if FIELD_GPS in fields:
            headers.append(COL_LAT)
            headers.append(COL_LON)
        if FIELD_PERSONS in fields:
            if person_mode == PERSON_MODE_LIST:
                headers.append(COL_PERSONS)
            else:
                for i in range(1, max(max_persons, 1) + 1):
                    headers.append(f"Person{i}")
        return headers


def _with_suffix(target_path: str | Path, suffix: str) -> Path:
    path = Path(target_path)
    if path.suffix.lower() == suffix:
        return path
    if path.suffix:
        return path.with_suffix(suffix)
    return path.with_suffix(suffix)
